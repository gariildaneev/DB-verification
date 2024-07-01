import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
import re
import tkinter as tk
from tkinter import filedialog, messagebox

def contains_cyrillic(text):
    return bool(re.search('[а-яА-Я]', text))

def highlight_cyrillic(cell):
    text = cell.value
    if not text:
        return

    parts = re.split(r'([а-яА-Я])', text)  # Разбить текст на части, включая кириллические символы
    new_text = ""
    for part in parts:
        if contains_cyrillic(part):
            new_text += f"<cyrillic>{part}</cyrillic>"
        else:
            new_text += part
    
    # Заменяем содержимое ячейки на новое
    cell.value = new_text
    
    # Применяем стили
    run = cell._element.xpath(".//r")
    for r in run:
        if "<cyrillic>" in r.text:
            r.text = r.text.replace("<cyrillic>", "")
            r.text = r.text.replace("</cyrillic>", "")
            rPr = r.get_or_add_rPr()
            rPr.append(PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"))  # Выделить желтым цветом

def validate_kks(input_file, output_file, check_unique, check_cyrillic):
    df = pd.read_excel(input_file)
    wb = Workbook()

    if check_unique:
        duplicates = df[df.duplicated(subset=['KKS'], keep=False)]
        ws_duplicates = wb.active
        ws_duplicates.title = "Отчет о дубликатах"
        ws_duplicates['A1'] = "Значение KKS не уникально"
        ws_duplicates['A1'].font = Font(size=14, bold=True)

        for r_idx, r in enumerate(dataframe_to_rows(duplicates, index=False, header=True), start=3):
            for c_idx, value in enumerate(r, start=1):
                ws_duplicates.cell(row=r_idx, column=c_idx, value=value)
    else:
        wb.remove(wb.active)

    if check_cyrillic:
        cyrillic_rows = df[df['KKS'].apply(contains_cyrillic)]
        ws_cyrillic = wb.create_sheet(title="Отчет о кириллице")
        ws_cyrillic['A1'] = "Значение KKS содержит кириллицу"
        ws_cyrillic['A1'].font = Font(size=14, bold=True)

        for r_idx, r in enumerate(dataframe_to_rows(cyrillic_rows, index=False, header=True), start=3):
            for c_idx, value in enumerate(r, start=1):
                cell = ws_cyrillic.cell(row=r_idx, column=c_idx, value=value)
                if c_idx == 3:  # Предполагается, что колонка 'KKS' первая
                    highlight_cyrillic(cell)

    wb.save(output_file)

def select_file():
    input_file = filedialog.askopenfilename(
        title="Выберите файл",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if input_file:
        output_file = filedialog.asksaveasfilename(
            title="Сохранить отчет как",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if output_file:
            try:
                validate_kks(input_file, output_file, check_unique.get(), check_cyrillic.get())
                messagebox.showinfo("Успех", "Отчет успешно создан!")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Произошла ошибка: {e}")

def create_gui():
    root = tk.Tk()
    root.title("Проверка базы данных")
    root.geometry("300x200")

    # Установка иконки для основного окна
    root.iconbitmap('./_internal/assets/icon.ico')

    global check_unique, check_cyrillic
    check_unique = tk.BooleanVar()
    check_cyrillic = tk.BooleanVar()

    chk_unique = tk.Checkbutton(root, text="Проверка уникальности KKS", variable=check_unique)
    chk_unique.pack(pady=5)

    chk_cyrillic = tk.Checkbutton(root, text="Проверка на кириллицу", variable=check_cyrillic)
    chk_cyrillic.pack(pady=5)

    btn_select_file = tk.Button(root, text="Выбрать файл", command=select_file)
    btn_select_file.pack(pady=20)

    root.mainloop()

if __name__ == "__main__":
    create_gui()
